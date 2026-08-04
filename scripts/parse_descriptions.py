# -*- coding: utf-8 -*-
"""統合Excel → src/data/descriptions.json 変換スクリプト。

使い方:
    python scripts/parse_descriptions.py

入力 (正本): 資料/クラスデータ/
  - 黒い砂漠M_スキル・パッシブ.xlsx
      「パッシブ一覧」 … 横=クラス / 縦=①②。セルは
                         「[日本語 / 韓国語 / 英語]」のヘッダー行 + 本文
      「スキル_{CODE}」 … 1行=1スキル。名前(日本語)/区分/STACK/CT/HIT/説明/種別 を読む
                         種別=闇精霊の怒りの行は rage として扱い、説明中の
                         「[武器分岐N]」マーカーで共通部と武器別2区間に再分割する
  - 黒い砂漠M_プロフィール.xlsx
      「プロフィール_{CODE}」 … A列ラベル(名前/出身地/Other/エピソード) B列に値

表示ルール (2026-07-15 まな指示 / 2026-08-03 更新):
  - 改行・字下げ・「・」・効果の順番は変えない (行をそのまま配列に入れる)
  - 秘伝の種別タグ(系列/無欠/…)とPvE/PvP行はExcel生成時に構造化済み
  - 最大使用回数/再使用待機時間/最大〇打撃は STACK / CT / HIT の数値として分離済み
    (HP側では「STACK:2　CT:4　HIT×2」の形式で表示する)
"""
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent  # skill-sigil-web/
DATA_DIR = ROOT / "資料" / "クラスデータ"
SRC_SKILLS = DATA_DIR / "黒い砂漠M_スキル・パッシブ.xlsx"
SRC_PROFILE = DATA_DIR / "黒い砂漠M_プロフィール.xlsx"
OUT = ROOT / "src" / "data" / "descriptions.json"
MASTER = ROOT / "src" / "data" / "master.json"

# 「スキル_{CODE}」シートの列 (1始まり)
COL_SKILL_ID = 1
COL_KIND = 2
COL_DISPLAY_NO = 3
COL_NAME_JA = 4
COL_REGION = 5
COL_STACK = 6
COL_CT = 7
COL_HIT = 8
COL_DESC = 9

KIND_NORMAL = "スキル"
KIND_RABAM = "ラバム"
KIND_RAGE = "闇精霊の怒り"

PASSIVE_HEAD = re.compile(r"^\[(.+?)\s*/\s*(.+?)\s*/\s*(.+?)\]$")
WEAPON_MARK = re.compile(r"^\[武器分岐(\d+)\]$")

warnings = []


def cell_lines(value):
    """セルの複数行テキストを行配列に。空行は落とす (元の字下げは保持)。"""
    if not isinstance(value, str):
        return []
    return [ln for ln in value.split("\n") if ln.strip() != ""]


def parse_region(value):
    """区分セル ("PvE" / "PvP" / "PvE・PvP") → (pve, pvp)"""
    s = str(value or "")
    return ("PvE" in s, "PvP" in s)


def to_int(value):
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def to_str(value):
    if value is None or str(value).strip() == "":
        return None
    return str(value).strip()


def parse_passives(ws, classes):
    """「パッシブ一覧」(横=クラス/縦=①②) → {code: [passive, passive]}"""
    col_of = {}
    for c in range(2, ws.max_column + 1):
        code = ws.cell(row=1, column=c).value
        if code:
            col_of[str(code).strip()] = c

    out = {}
    for code in classes:
        col = col_of.get(code)
        if col is None:
            warnings.append(f"{code}: パッシブ一覧に列なし")
            out[code] = []
            continue
        passives = []
        for row in (2, 3):  # パッシブ①/②
            lines = cell_lines(ws.cell(row=row, column=col).value)
            name = ""
            if lines:
                m = PASSIVE_HEAD.match(lines[0].strip())
                if m:
                    name = m.group(1).strip()
                    lines = lines[1:]
                else:
                    warnings.append(f"{code}: パッシブ{row - 1} の名称ヘッダーがありません")
            passives.append({"name": name, "lines": lines})
        out[code] = passives
    return out


def split_rage(lines):
    """rage本文を [武器分岐N] マーカーで 共通部 + 武器別(0 or 2区間) に分割。"""
    marks = [i for i, ln in enumerate(lines) if WEAPON_MARK.match(ln.strip())]
    if not marks:
        return lines, []
    common = lines[: marks[0]]
    sections = []
    for i, start in enumerate(marks):
        end = marks[i + 1] if i + 1 < len(marks) else len(lines)
        sections.append(lines[start + 1: end])  # マーカー行自体は出力しない
    if len(sections) == 1:
        sections.append([])
    elif len(sections) > 2:
        warnings.append(f"武器分岐が3区間以上あります ({len(sections)})")
        sections = sections[:2]
    return common, sections


def parse_skill_sheet(ws, code):
    """1クラスのスキルシート → (skills{}, rage)"""
    skills = {}
    rage = None
    for r in range(2, ws.max_row + 1):
        skill_id = ws.cell(row=r, column=COL_SKILL_ID).value
        kind = str(ws.cell(row=r, column=COL_KIND).value or "").strip()
        if not skill_id or not kind:
            continue

        name = str(ws.cell(row=r, column=COL_NAME_JA).value or "").strip()
        pve, pvp = parse_region(ws.cell(row=r, column=COL_REGION).value)
        lines = cell_lines(ws.cell(row=r, column=COL_DESC).value)

        if kind == KIND_RAGE:
            common, weapon = split_rage(lines)
            rage = {
                "name": name,
                "pve": pve,
                "pvp": pvp,
                "common": common,
                "weapon": weapon,
            }
            continue

        if kind not in (KIND_NORMAL, KIND_RABAM):
            warnings.append(f"{code}: 未知の種別 '{kind}' ({skill_id})")
            continue

        display_no = to_int(ws.cell(row=r, column=COL_DISPLAY_NO).value)
        if display_no is None:
            warnings.append(f"{code}: {skill_id} の表示番号が読めません")
            continue

        key = f"{'sp' if kind == KIND_RABAM else 'n'}_{display_no}"
        skills[key] = {
            "name": name,
            "pve": pve,
            "pvp": pvp,
            "rabam": kind == KIND_RABAM,
            "stack": to_int(ws.cell(row=r, column=COL_STACK).value),
            "ct": to_str(ws.cell(row=r, column=COL_CT).value),
            "hit": to_int(ws.cell(row=r, column=COL_HIT).value),
            "lines": lines,
        }

    if rage is None:
        warnings.append(f"{code}: 闇精霊の怒りの行がありません")
    return skills, rage


def parse_profiles(classes):
    """プロフィールExcel → {code: {name, hometown, other, episode}}"""
    wb = openpyxl.load_workbook(SRC_PROFILE, read_only=True, data_only=True)
    sheet_by_name = {n.strip(): n for n in wb.sheetnames}
    out = {}
    for code in classes:
        label = f"プロフィール_{code}"
        if label not in sheet_by_name:
            warnings.append(f"{code}: {label} シートなし")
            out[code] = {"name": "", "hometown": "", "other": None, "episode": None}
            continue
        ws = wb[sheet_by_name[label]]
        values = {}
        for r in range(1, ws.max_row + 1):
            k = ws.cell(row=r, column=1).value
            if k:
                values[str(k).strip()] = to_str(ws.cell(row=r, column=2).value)
        if not values.get("名前"):
            warnings.append(f"{code}: プロフィールの名前が空欄")
        out[code] = {
            "name": values.get("名前") or "",
            "hometown": values.get("出身地") or "",
            "other": values.get("Other"),
            "episode": values.get("エピソード"),
        }
    wb.close()
    return out


def main():
    master = json.loads(MASTER.read_text(encoding="utf-8"))
    codes = [c["code"] for c in master["classes"]]

    wb = openpyxl.load_workbook(SRC_SKILLS, read_only=True, data_only=True)
    sheet_by_name = {n.strip(): n for n in wb.sheetnames}

    if "パッシブ一覧" not in sheet_by_name:
        raise SystemExit(f"「パッシブ一覧」シートがありません: {SRC_SKILLS}")
    passives_by_class = parse_passives(wb[sheet_by_name["パッシブ一覧"]], codes)
    profiles = parse_profiles(codes)

    classes = {}
    for code in codes:
        label = f"スキル_{code}"
        if label not in sheet_by_name:
            warnings.append(f"{code}: {label} シートなし")
            continue
        skills, rage = parse_skill_sheet(wb[sheet_by_name[label]], code)

        # master.json の名称と照合 (統合Excelはmaster準拠のはずなので不一致は要調査)
        for s in master["skills"].get(code, []):
            key = f"{'sp' if s['group'] == 'special' else 'n'}_{s['display_no']}"
            d = skills.get(key)
            if d and d["name"] and s["name_ja"] and d["name"] != s["name_ja"]:
                warnings.append(
                    f"{code}/{key}: 名称不一致 master={s['name_ja']!r} excel={d['name']!r}"
                )

        classes[code] = {
            "passives": passives_by_class.get(code, []),
            "rage": rage,
            "skills": skills,
            "profile": profiles[code],
        }

    wb.close()

    data = {"schema_version": 2, "classes": classes}
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=1) + "\n", encoding="utf-8")
    size = OUT.stat().st_size
    print(f"OK: {len(classes)}クラス → {OUT} ({size / 1024:.0f} KB)")
    if warnings:
        print(f"--- 警告 {len(warnings)}件 ---")
        for w in warnings:
            print(" ", w)


if __name__ == "__main__":
    main()
