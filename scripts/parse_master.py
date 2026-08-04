# -*- coding: utf-8 -*-
"""統合Excel → src/data/master.json 変換スクリプト。

使い方:
    python scripts/parse_master.py [統合Excelのパス]

入力 (正本): 資料/クラスデータ/黒い砂漠M_スキル・パッシブ.xlsx
  - 「クラス一覧」シート … コード / クラス名 / 表示順 / 有効
  - 「スキル_{CODE}」シート … 1行=1スキル。ここから skill_id / 種別 / 表示番号 /
    名前(日本語) / 装着可能秘伝 を読む (種別=闇精霊の怒りの行は master 対象外)

秘伝の種類・等級・効果は src/game-rules/skill-sigil.json が正本。
アプリは src/data/master.ts でそれと合成するため、ここでは出力しない。
"""
import json
import sys
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent  # skill-sigil-web/
DEFAULT_SRC = ROOT / "資料" / "クラスデータ" / "黒い砂漠M_スキル・パッシブ.xlsx"
GAME_RULES = ROOT / "src" / "game-rules" / "skill-sigil.json"
OUT = ROOT / "src" / "data" / "master.json"

# マスタの中身(クラス・スキル・固定枠)が変わったときだけ上げる。入力Excelを
# 統合版に移行しただけでは中身は変わらないため据え置く (上げるとユーザーの
# バックアップ復元時に「マスタ版が異なります」の警告が出る。src/logic/excel.ts)
MASTER_VERSION = "2026-07-13"

# 「スキル_{CODE}」シートの列 (1始まり)
COL_SKILL_ID = 1
COL_KIND = 2
COL_DISPLAY_NO = 3
COL_NAME_JA = 4
COL_SLOTS = 11

KIND_TO_GROUP = {"スキル": "normal", "ラバム": "special"}


def norm(s):
    if s is None:
        return ""
    return unicodedata.normalize("NFKC", str(s)).strip()


def load_slot_ids():
    """秘伝タイプの日本語表記 → id の逆引き辞書 (name と alias の両方を受ける)。"""
    rules = json.loads(GAME_RULES.read_text(encoding="utf-8"))
    table = {}
    for t in rules["sigil_types"]:
        table[t["name"]] = t["id"]
        if t.get("alias"):
            table[t["alias"]] = t["id"]
    return table


def parse_classes(ws, warnings):
    classes = []
    for r in range(2, ws.max_row + 1):
        code = norm(ws.cell(row=r, column=1).value)
        if not code:
            continue
        name = norm(ws.cell(row=r, column=2).value)
        order_raw = ws.cell(row=r, column=3).value
        enabled = ws.cell(row=r, column=4).value
        try:
            sort_order = int(order_raw)
        except (TypeError, ValueError):
            sort_order = len(classes)
            warnings.append(f"{code}: 表示順が数値でないため {sort_order} を補完")
        if not name:
            warnings.append(f"{code}: クラス名が空欄")
        classes.append({
            "class_id": code,
            "code": code,
            "name_ja": name or code,
            "sort_order": sort_order,
            "enabled": bool(enabled),
        })
    classes.sort(key=lambda c: c["sort_order"])
    return classes


def parse_skill_sheet(ws, code, slot_ids, warnings):
    """1クラスのスキルシートから通常13 + ラバム4を抽出する。"""
    skills = []
    for r in range(2, ws.max_row + 1):
        skill_id = norm(ws.cell(row=r, column=COL_SKILL_ID).value)
        kind = norm(ws.cell(row=r, column=COL_KIND).value)
        if not skill_id or kind not in KIND_TO_GROUP:
            continue  # 空行 / 闇精霊の怒り

        try:
            display_no = int(float(norm(ws.cell(row=r, column=COL_DISPLAY_NO).value)))
        except ValueError:
            warnings.append(f"{code}: {skill_id} の表示番号が読めません (row {r})")
            continue

        name = norm(ws.cell(row=r, column=COL_NAME_JA).value)
        if not name:
            warnings.append(f"{code}: {skill_id} の名前が空欄")

        slots = []
        raw_slots = norm(ws.cell(row=r, column=COL_SLOTS).value)
        if raw_slots:
            for label in raw_slots.split("・"):
                label = label.strip()
                if not label:
                    continue
                sid = slot_ids.get(label)
                if sid is None:
                    warnings.append(f"{code}: 不明な秘伝タイプ '{label}' ({skill_id})")
                else:
                    slots.append(sid)
            if len(slots) not in (0, 4):
                warnings.append(f"{code}: {skill_id} の枠数が4でない ({len(slots)})")

        skills.append({
            "skill_id": skill_id,
            "group": KIND_TO_GROUP[kind],
            "display_no": display_no,
            "name_ja": name or f"(名称未設定 {display_no})",
            "sigil_eligible": len(slots) == 4,
            "slots": slots if len(slots) == 4 else None,
        })

    return sorted(skills, key=lambda s: (0 if s["group"] == "special" else 1, s["display_no"]))


def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SRC
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    sheet_by_name = {n.strip(): n for n in wb.sheetnames}

    warnings = []
    slot_ids = load_slot_ids()

    if "クラス一覧" not in sheet_by_name:
        raise SystemExit(f"「クラス一覧」シートがありません: {src}")
    classes = parse_classes(wb[sheet_by_name["クラス一覧"]], warnings)

    skills_by_class = {}
    for cls in classes:
        code = cls["code"]
        label = f"スキル_{code}"
        if label not in sheet_by_name:
            warnings.append(f"{code}: {label} シートなし")
            continue
        skills = parse_skill_sheet(wb[sheet_by_name[label]], code, slot_ids, warnings)
        n_special = sum(1 for s in skills if s["group"] == "special")
        n_normal = sum(1 for s in skills if s["group"] == "normal")
        if n_special != 4 or n_normal != 13:
            warnings.append(f"{code}: スキル数異常 special={n_special} normal={n_normal}")
        skills_by_class[code] = skills

    wb.close()

    master = {
        "schema_version": 1,
        "master_version": MASTER_VERSION,
        "classes": classes,
        "skills": skills_by_class,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(master, ensure_ascii=False, indent=1) + "\n", encoding="utf-8")

    total = sum(len(v) for v in skills_by_class.values())
    print(f"classes={len(classes)} skills={total} -> {OUT}")
    for w in warnings:
        print("WARN:", w)


if __name__ == "__main__":
    main()
